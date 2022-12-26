import openpyxl
from googletrans import Translator
from aiogram import types
import pytesseract
import cv2
import os

from filters.all_filters import IsGroup
from loader import dp, database


@dp.message_handler(IsGroup(), content_types=[types.ContentType.PHOTO, types.ContentType.DOCUMENT])
async def translate_image(message: types.Message):
    """Распознаёт только .png картинки и только на английском языке"""
    file_path = f'data/{message.from_user.id}.png'  # Создаём путь, по которому сохранится фото
    await message.photo.pop().download(destination_file=file_path)  # Сохраняем фото по нужному пути
    image = cv2.imread(file_path)
    text = pytesseract.image_to_string(image)
    translated_texts = []
    translator = Translator()
    user_languages = {line[1] for line in database.select_all()}
    workbook = openpyxl.load_workbook('data/available_languages.xlsx')
    worksheet = workbook.active
    for user_lang in user_languages:
        for i in range(worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column + 1):
                if col[i].value == user_lang:
                    language = worksheet[f'A{i + 1}'].value
                    break
        translated_texts.append(translator.translate(text=f'{language}: ' + text, dest=user_lang).text)
    await message.answer('\n\n'.join(translated_texts))
    # except:
    #     await message.answer('Не получилось прочитать картинку. Возможно, вы отправили не .png файл')
    os.remove(file_path)  # Удаление временного файла

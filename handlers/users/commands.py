import sqlite3
import openpyxl
from aiogram import types
from aiogram.dispatcher.filters.builtin import CommandHelp, CommandStart, Command
from translate_text import translate

from loader import dp, database
from utils.misc import rate_limit


@rate_limit(5, 'help')
@dp.message_handler(CommandHelp())
async def bot_help(message: types.Message):
    text = [
        'Список команд: ',
        '/help - Получить справку',
        '/start - Перезапуск бота',
        '/showalllangs - Показать все доступные языки',
        '/setmylang ... - Задать мой язык'
    ]
    await message.answer('\n'.join(text))


@dp.message_handler(CommandStart())
async def bot_start(message: types.Message):
    user_id = message.from_user.id
    msg = translate(f'Привет, {message.from_user.full_name}!', user_id=user_id)
    await message.answer(msg)
    try:
        database.add_user(id=user_id, language=message.from_user.language_code)
    except sqlite3.IntegrityError:
        pass


@dp.message_handler(Command('showalllangs'))
async def show_all_languages(message: types.Message):
    await message.answer("""
    Выбери свой язык, на который мне нужно будет переводить сообщения
    И отправь мне команду с выбранным языком
    Примеры: 
    /setmylang en
    /setmylang eng
    /setmylang 45""")
    await dp.bot.send_document(message.chat.id, open('data/available_languages.txt', 'rb'))


@dp.message_handler(Command('setmylang'))
async def set_user_language(message: types.Message):
    language_code = message.text.replace('/setmylang ', '')
    try:
        workbook = openpyxl.load_workbook('data/available_languages.xlsx')
        worksheet = workbook.active
        for i in range(worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column+1):
                if col[i].value == language_code:
                    language_code = worksheet[f'B{i+1}'].value
                    language = worksheet[f'A{i+1}'].value
                    break
        await message.answer(f'Отлично, теперь твой язык {language.lower()}')
        database.update_language(id=message.from_user.id, new_language=language_code)
    except:
        await message.answer(f'Некорректный код языка: {language_code}.\n'
                             f'Посмотрите таблицу со всеми поддерживаемыми языками с помощью команды /showalllangs')


@dp.message_handler(Command('check'))
async def check(message: types.Message):
    await message.answer(f'{database.select_all()}')

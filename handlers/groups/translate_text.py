import openpyxl
from aiogram import types
from googletrans import Translator

from filters.all_filters import IsGroup
from loader import dp, database


@dp.message_handler(IsGroup())
async def translate(message: types.Message):
    translated_texts = []
    translator = Translator()
    user_languages = {line[1] for line in database.select_all()}
    workbook = openpyxl.load_workbook('data/available_languages.xlsx')
    worksheet = workbook.active
    for user_lang in user_languages:
        for i in range(worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column + 1):
                if col[i].value == user_lang:
                    language = worksheet[f'A{i + 1}'].value
                    break
        translated_texts.append(translator.translate(text=f'{language}: ' + message.text, dest=user_lang).text)
    await message.answer('\n\n'.join(translated_texts))

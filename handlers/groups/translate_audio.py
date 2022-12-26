import os
from os import path
import soundfile
import openpyxl
import requests
import speech_recognition
from aiogram import types
from googletrans import Translator

from data.config import BOT_TOKEN
from filters.all_filters import IsGroup
from loader import dp, bot, database


@dp.message_handler(IsGroup(), content_types=[
    types.ContentType.VOICE,
    types.ContentType.AUDIO
])
async def translate_audio(message: types.Message):
    """Работает только с .wav расширением и распознаёт только английский язык"""
    # TODO не работает: soundfile.LibsndfileError: Supported file format but file is malformed.
    # Получаем информацию о файле
    # if message.content_type == types.ContentType.VOICE:
    #     file_info = await bot.get_file(message.voice.file_id)
    #     file = requests.get(f'https://api.telegram.org/file/bot{BOT_TOKEN}/{file_info.file_path}')
    #     file_path = path.join(os.getcwd(), 'data', f'{message.from_user.id}.ogg')
    #     with open(file_path, 'wb') as f:
    #         f.write(file.content)
    #     data, sample_rate = soundfile.read(f'data/{message.from_user.id}.ogg')
    #     soundfile.write(file_path, data, sample_rate)
    #     file_path = path.join(os.getcwd(), 'data', f'{message.from_user.id}.wav')
    # else:
    file_info = await bot.get_file(message.audio.file_id)
    # Скачиваем файл себе на компьютер
    file = requests.get(f'https://api.telegram.org/file/bot{BOT_TOKEN}/{file_info.file_path}')
    file_path = path.join(os.getcwd(), 'data', f'{message.from_user.id}.wav')
    with open(file_path, 'wb') as f:
        f.write(file.content)

    # Распознаём речь и переводим её в текст
    recognizer = speech_recognition.Recognizer()
    with speech_recognition.AudioFile(file_path) as source:
        audio = recognizer.record(source)
    text = recognizer.recognize_google(audio, language='en-EN')

    # Отправляем текст пользователю
    translated_texts = []
    translator = Translator()
    user_languages = {line[1] for line in database.select_all()}
    workbook = openpyxl.load_workbook('data/available_languages.xlsx')
    worksheet = workbook.active
    for user_lang in user_languages:
        for i in range(worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column + 1):
                if col[i].value == user_lang:
                    language = worksheet[f'A{i + 1}'].value
                    break
        translated_texts.append(translator.translate(text=f'{language}: ' + text, dest=user_lang).text)
    await message.answer('\n\n'.join(translated_texts))

    os.remove(file_path)  # Удаление временного файла
